import os
import glob
from collections import defaultdict
from xml.etree import ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NS = {"n": "http://www.sped.fazenda.gov.br/nfse"}

CNPJS_ELECNOR = {
    "30455661000172": "Elecnor Brasil - Matriz",
    "30455661000920": "Elecnor Brasil - Osório",
    "30455661001659": "Elecnor Brasil - Paracatu",
    "30455661001730": "Elecnor Brasil - Estiva Gerbi",
    "30455661001810": "Elecnor Brasil - Obidos",
    "30455661001900": "Elecnor Brasil - Barreiras",
    "30455661002035": "Elecnor Brasil - Miracema do Tocantins",
    "30455661002116": "Elecnor Brasil - Piripiri",
    "30455661002205": "Elecnor Brasil - Tianguá",
    "30455661002388": "Elecnor Brasil - Lajes",
    "30455661002469": "Elecnor Brasil - Manaus",
    "30455661002540": "Elecnor Brasil - Coremas",
    "30455661002620": "Elecnor Brasil - Bernardo do Mearim",
    "30455661002701": "Elecnor Brasil - Nova Alvorada do Sul",
    "30455661002892": "Elecnor Brasil - Rio das Ostras",
    "30455661002973": "Elecnor Brasil - Araripina",
    "30455661003007": "Elecnor Brasil - Arinos",
    "30455661003198": "Elecnor Brasil - Tobias Barreto",
    "30455661003279": "Elecnor Brasil - Olimpia",
    "30455661003350": "Elecnor Brasil - Canindé de São Francisco",
    # Azulão — certificado separado
    "44987240000105": "Elecnor Azulão - Matriz",
    "44987240000288": "Elecnor Azulão - Filial",
}

CSTAT = {
    "100": "Gerada (ADN)",
    "101": "Gerada em Substituição",
    "102": "Decisão Judicial/Adm.",
    "103": "Avulsa",
    "107": "Gerada (Município)",
    "108": "Cancelada (Município)",
}

MOTIVO_CANCEL = {
    "1": "Erro na emissão",
    "2": "Serviço não concluído",
    "3": "Duplicidade de NFS-e",
    "4": "Erro de tributação",
    "99": "Outros",
}

TIPO_EVENTO_TAG = {
    "e101101": "Cancelamento",
    "e101102": "Cancelamento (decisão judicial)",
        "e105102": "Cancelamento por substituição",
    "e102101": "Confirmação pelo prestador",
    "e102102": "Rejeição pelo prestador",
    "e103101": "Confirmação pelo tomador",
    "e103102": "Rejeição pelo tomador",
    "e104101": "Confirmação pelo intermediário",
    "e104102": "Rejeição pelo intermediário",
    "e106101": "Confirmação tácita",
    "e107101": "Anulação de rejeição",
    "e108101": "Cancelamento por ofício",
    "e109101": "Bloqueio por ofício",
    "e110101": "Desbloqueio por ofício",
}

MESES_PT = {
    "01": "Janeiro", "02": "Fevereiro", "03": "Março",
    "04": "Abril",   "05": "Maio",      "06": "Junho",
    "07": "Julho",   "08": "Agosto",    "09": "Setembro",
    "10": "Outubro", "11": "Novembro",  "12": "Dezembro",
}

UNIDADE_DESCONHECIDA = "Não Identificada"

AZUL       = "003F87"
AZUL_CLARO = "ECF5FE"
LARANJA    = "F47C00"
VERDE      = "1A6B3C"
VERDE_CLAR = "E8F5EE"
ROXO       = "4B2D8F"
ROXO_CLARO = "F0EBFA"
BRANCO     = "FFFFFF"
CINZA      = "F5F8FB"

# ── helpers ────────────────────────────────────────────────
def limpar_cnpj(v):
    return "".join(filter(str.isdigit, str(v or "")))

def formatar_cnpj(cnpj):
    c = limpar_cnpj(cnpj)
    return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}" if len(c) == 14 else cnpj

def identificar_unidade(cnpj):
    return CNPJS_ELECNOR.get(limpar_cnpj(cnpj), "")

def is_elecnor(cnpj):
    return limpar_cnpj(cnpj) in CNPJS_ELECNOR

def derivar_papel_elecnor(cnpj_prestador: str, cnpj_tomador: str,
                          cnpj_intermediario: str = "") -> str:
    """
    Determina o papel da Elecnor na nota:
    Prestadora / Tomadora / Intermediária / Intragrupo / Indefinido
    """
    prest  = limpar_cnpj(cnpj_prestador)
    tomad  = limpar_cnpj(cnpj_tomador)
    interm = limpar_cnpj(cnpj_intermediario)
    eh_prest  = prest  in CNPJS_ELECNOR
    eh_tomad  = tomad  in CNPJS_ELECNOR
    eh_interm = interm in CNPJS_ELECNOR
    if eh_prest and eh_tomad:
        return "Intragrupo"
    if eh_prest:
        return "Prestadora"
    if eh_tomad:
        return "Tomadora"
    if eh_interm:
        return "Intermediária"
    return "Indefinido"

def txt(el, tag):
    if el is None: return ""
    f = el.find(f"n:{tag}", NS)
    return (f.text or "").strip() if f is not None else ""

def to_float(val):
    try: return float((val or "").replace(",", "."))
    except: return None

def competencia_chave(data_str):
    if data_str and len(data_str) >= 7:
        return data_str[:7]
    return "0000-00"

def competencia_label(chave):
    try:
        ano, mes = chave.split("-")
        return f"{MESES_PT.get(mes, mes)}/{ano}"
    except:
        return chave

def slug_seguro(texto):
    """Remove caracteres inválidos para nomes de pasta/arquivo."""
    invalidos = r'\/:*?"<>|'
    for c in invalidos:
        texto = texto.replace(c, "_")
    return texto.strip()

# ── parsers ────────────────────────────────────────────────
def parse_nfse(path):
    tree = ET.parse(path)
    root = tree.getroot()
    inf = root.find("n:infNFSe", NS)
    if inf is None: return None

    emit      = inf.find("n:emit", NS)
    cnpj_emit = txt(emit, "CNPJ")
    nome_emit = txt(emit, "xNome")
    im_emit   = txt(emit, "IM")
    uf_emit   = txt(emit.find("n:enderNac", NS) if emit is not None else None, "UF")

    # ── Valores do nível infNFSe (calculados pelo ADN) ──────────
    # Fonte correta conforme leiaute: infNFSe/valores
    # vBC, pAliqAplic, vISSQN e vLiq ficam aqui — não em infDPS/valores
    vals_inf = inf.find("n:valores", NS)
    vbc   = to_float(txt(vals_inf, "vBC"))         # base de cálculo do ISS (após deduções)
    paliq = to_float(txt(vals_inf, "pAliqAplic"))   # alíquota aplicada pelo ADN
    viss  = to_float(txt(vals_inf, "vISSQN"))       # valor do ISS calculado pelo ADN
    vliq  = to_float(txt(vals_inf, "vLiq"))         # valor líquido da NFS-e

    # Valor bruto do serviço vem de infDPS/valores/vServPrest/vServ
    # É lido depois junto com os demais tributos do DPS
    # ATENÇÃO: vServ ≠ vBC quando há deduções (vDedRed) — não substituir um pelo outro

    dps     = inf.find("n:DPS", NS)
    inf_dps = dps.find("n:infDPS", NS) if dps is not None else None

    dh_emi  = txt(inf_dps, "dhEmi")[:10] if inf_dps is not None else txt(inf, "dhProc")[:10]
    dcompet = txt(inf_dps, "dCompet") if inf_dps is not None else dh_emi

    toma       = inf_dps.find("n:toma", NS) if inf_dps is not None else None
    cnpj_toma  = (txt(toma, "CNPJ") or txt(toma, "CPF")) if toma is not None else ""
    nome_toma  = txt(toma, "xNome") if toma is not None else ""
    interm     = inf_dps.find("n:interm", NS) if inf_dps is not None else None
    cnpj_interm = (txt(interm, "CNPJ") or txt(interm, "CPF")) if interm is not None else ""

    serv  = inf_dps.find("n:serv", NS) if inf_dps is not None else None
    cserv = serv.find("n:cServ", NS) if serv is not None else None
    ctrib = txt(cserv, "cTribNac") if cserv is not None else ""
    xdesc = txt(cserv, "xDescServ") if cserv is not None else ""

    # ── Tributos ────────────────────────────────────────────
    vals_dps_el = inf_dps.find("n:valores", NS) if inf_dps is not None else None
    trib_el     = vals_dps_el.find("n:trib", NS) if vals_dps_el is not None else None
    trib_mun_el = trib_el.find("n:tribMun", NS) if trib_el is not None else None
    trib_fed_el = trib_el.find("n:tribFed", NS) if trib_el is not None else None
    piscofins   = trib_fed_el.find("n:piscofins", NS) if trib_fed_el is not None else None
    tot_trib_el = trib_el.find("n:totTrib", NS) if trib_el is not None else None
    vtot_el     = tot_trib_el.find("n:vTotTrib", NS) if tot_trib_el is not None else None

    # ISSQN
    TP_RET = {"1": "Não Retido", "2": "Retido pelo Tomador", "3": "Retido pelo Intermediário"}
    retencao_iss = TP_RET.get(txt(trib_mun_el, "tpRetISSQN"), "-")
    aliq_iss_dps = to_float(txt(trib_mun_el, "pAliq"))  # alíquota declarada no DPS

    # PIS / COFINS
    v_bc_piscofins  = to_float(txt(piscofins, "vBCPisCofins"))
    aliq_pis        = to_float(txt(piscofins, "pAliqPis"))
    aliq_cofins     = to_float(txt(piscofins, "pAliqCofins"))
    v_pis           = to_float(txt(piscofins, "vPis"))
    v_cofins        = to_float(txt(piscofins, "vCofins"))
    # tpRetPisCofins — define quais tributos do CSRF estão retidos
    # Conforme leiaute DPS/NFS-e: o campo indica o conjunto de contribuições retidas
    TP_RET_PISCOF = {
        "1": "Não Retido",
        "2": "PIS/COFINS/CSLL Retidos",
        "3": "PIS/COFINS/CSLL Retidos (Lei 9.711/98)",
        "4": "PIS/COFINS/CSLL Retidos (Lei 10.925/04)",
        "5": "PIS/COFINS/CSLL Retidos (Constr. Civil)",
        "6": "PIS/COFINS Retidos (Lei 10.833 Art.3)",
        "7": "PIS/COFINS Retidos (Lei 10.865/04)",
        "8": "PIS/COFINS/CSLL Retidos (Lei 11.196/05)",
        "9": "PIS/COFINS/CSLL Retidos (outros)",
    }
    tp_ret_raw    = txt(piscofins, "tpRetPisCofins")
    ret_piscofins = TP_RET_PISCOF.get(tp_ret_raw, "-")
    # Flags individuais derivadas do tpRetPisCofins para facilitar filtros
    csrf_retido   = "Sim" if tp_ret_raw not in ("", "1") else "Não"
    csll_retido   = "Sim" if tp_ret_raw in ("2","3","4","5","8","9") else "Não"

    # IRRF / INSS / CSLL
    v_irrf  = to_float(txt(trib_fed_el, "vRetIRRF"))
    v_inss  = to_float(txt(trib_fed_el, "vRetCP"))
    v_csll  = to_float(txt(trib_fed_el, "vRetCSLL"))

    # Totais aproximados
    v_tot_fed = to_float(txt(vtot_el, "vTotTribFed"))
    v_tot_est = to_float(txt(vtot_el, "vTotTribEst"))
    v_tot_mun = to_float(txt(vtot_el, "vTotTribMun"))

    # IBS / CBS (futuro — estrutura dentro de infNFSe/IBSCBS)
    ibscbs      = inf.find("n:IBSCBS", NS) if inf is not None else None
    ibscbs_vals = ibscbs.find("n:valores", NS) if ibscbs is not None else None
    tot_cibs    = ibscbs.find("n:totCIBS", NS) if ibscbs is not None else None
    g_ibs       = tot_cibs.find("n:gIBS", NS) if tot_cibs is not None else None
    g_cbs       = tot_cibs.find("n:gCBS", NS) if tot_cibs is not None else None
    v_ibs_tot   = to_float(txt(g_ibs, "vIBSTot"))
    v_cbs_tot   = to_float(txt(g_cbs, "vCBS"))
    v_tot_nf    = to_float(txt(tot_cibs, "vTotNF"))
    aliq_cbs    = to_float(txt(ibscbs_vals.find("n:fed", NS) if ibscbs_vals is not None else None, "pCBS"))

    # Valor do serviço
    vserv_p = vals_dps_el.find("n:vServPrest", NS) if vals_dps_el is not None else None
    vserv   = to_float(txt(vserv_p, "vServ")) if vserv_p is not None else None

    chave_nfse    = (txt(inf, "chNFSe") or inf.get("Id", "")).lstrip("NFS").lstrip("NFS")
    # Remove prefixo "NFS" do @Id para padronizar com a chave do evento
    if chave_nfse.startswith("NFS"):
        chave_nfse = chave_nfse[3:]
    situacao      = CSTAT.get(txt(inf, "cStat"), f"Cód {txt(inf, 'cStat')}")
    mun_incid     = txt(inf, "xLocIncid") or "-"
    cod_mun_incid = txt(inf, "cLocIncid") or "-"

    papel         = "Tomadora"      if is_elecnor(cnpj_toma)   else \
                    "Prestadora"    if is_elecnor(cnpj_emit)   else \
                    "Intermediária" if is_elecnor(cnpj_interm) else "Outro"
    papel_elecnor = derivar_papel_elecnor(cnpj_emit, cnpj_toma, cnpj_interm)
    unidade = identificar_unidade(cnpj_toma)   if papel == "Tomadora"      else \
              identificar_unidade(cnpj_emit)   if papel == "Prestadora"    else \
              identificar_unidade(cnpj_interm) if papel == "Intermediária"  else ""
    if not unidade:
        unidade = UNIDADE_DESCONHECIDA

    comp = competencia_chave(dcompet or dh_emi)

    return {
        "_papel":   papel,
        "_chave":   chave_nfse,
        "_comp":    comp,
        "_unidade": unidade,
        "Arquivo":           os.path.basename(path),
        "Chave NFS-e":       chave_nfse,
        "Número NFS-e":      txt(inf, "nNFSe"),
        "Data Emissão":      dh_emi,
        "Competência":       dcompet or dh_emi,
        "Situação":          situacao,
        "Unidade Elecnor":   unidade,
        "Papel Elecnor":     papel_elecnor,
        "CNPJ Prestador":    formatar_cnpj(cnpj_emit),
        "Nome Prestador":    nome_emit,
        "IM Prestador":      im_emit,
        "UF Prestador":      uf_emit,
        "CNPJ/CPF Tomador":  formatar_cnpj(cnpj_toma),
        "Nome Tomador":      nome_toma,
        "Cód. Serviço":      ctrib,
        "Descrição":         xdesc,
        "Valor Serviço":     vserv,
        "Base de Cálculo":   vbc,
        "Alíquota ISS %":    paliq,
        "Valor ISS":         viss,
        "Valor Líquido":     vliq,
        # ISSQN
        # Localidade de incidência do ISS
        "Município Incidência ISS": mun_incid,
        "Cód. IBGE Incidência":     cod_mun_incid,
        "Retenção ISS":      retencao_iss,
        "Alíq. ISS DPS":     aliq_iss_dps,
        # PIS / COFINS
        "BC PIS/COFINS":     v_bc_piscofins,
        "Alíq. PIS %":       aliq_pis,
        "Alíq. COFINS %":    aliq_cofins,
        "Valor PIS":         v_pis,
        "Valor COFINS":      v_cofins,
        "Retenção PIS/COF":  ret_piscofins,
        "CSRF Retido":       csrf_retido,
        "CSLL Retida":       csll_retido,
        # Federais retidos
        "Valor IRRF":        v_irrf,
        "Valor INSS":        v_inss,
        "Valor CSLL":        v_csll,
        # Totais aproximados
        "Tot. Trib. Fed.":   v_tot_fed,
        "Tot. Trib. Est.":   v_tot_est,
        "Tot. Trib. Mun.":   v_tot_mun,
        # IBS / CBS (a partir de 2027)
        "Valor IBS":         v_ibs_tot,
        "Valor CBS":         v_cbs_tot,
        "Alíq. CBS %":       aliq_cbs,
        "Vlr Líq. + IBS/CBS": v_tot_nf,
        "Tem Evento":        "",
        "Tipo Evento":       "",
        "Data Evento":       "",
        "Motivo Evento":     "",
    }

def parse_evento(path):
    tree = ET.parse(path)
    root = tree.getroot()
    inf     = root.find("n:infEvento", NS)
    if inf is None: return None
    ped     = inf.find("n:pedRegEvento", NS)
    inf_ped = ped.find("n:infPedReg", NS) if ped is not None else None
    if inf_ped is None: return None

    tipo_tag = xdesc_ev = cmotivo = xmotivo = ""
    for child in inf_ped:
        tag = child.tag.split("}")[-1]
        if tag.startswith("e") and len(tag) > 1:
            tipo_tag = tag
            xd = child.find("n:xDesc", NS)
            cm = child.find("n:cMotivo", NS)
            xm = child.find("n:xMotivo", NS)
            xdesc_ev = xd.text if xd is not None else ""
            cmotivo  = cm.text if cm is not None else ""
            xmotivo  = xm.text if xm is not None else ""
            break

    tipo_legivel   = TIPO_EVENTO_TAG.get(tipo_tag, xdesc_ev or tipo_tag)
    motivo_legivel = MOTIVO_CANCEL.get(cmotivo, xmotivo or cmotivo)
    cnpj_autor     = txt(inf_ped, "CNPJAutor")
    dh_evento      = txt(inf_ped, "dhEvento")[:10]
    unidade        = identificar_unidade(cnpj_autor) or UNIDADE_DESCONHECIDA

    return {
        "_chave_nfse": txt(inf_ped, "chNFSe"),
        "_comp":       competencia_chave(dh_evento),
        "_unidade":    unidade,
        "Arquivo":             os.path.basename(path),
        "Chave NFS-e":         txt(inf_ped, "chNFSe"),
        "Tipo de Evento":      tipo_legivel,
        "Data do Evento":      dh_evento,
        "Data Processamento":  txt(inf, "dhProc")[:10],
        "Nº Seq. Evento":      txt(inf, "nSeqEvento"),
        "CNPJ Autor":          formatar_cnpj(cnpj_autor),
        "Unidade Elecnor":     unidade,
        "Cód. Motivo":         cmotivo,
        "Motivo":              motivo_legivel,
        "_tag":                tipo_tag,
    }

# ── colunas ────────────────────────────────────────────────
COLS_NFSE = [
    "Arquivo", "Chave NFS-e", "Número NFS-e", "Data Emissão", "Competência",
    "Situação", "Unidade Elecnor", "Papel Elecnor",
    "CNPJ Prestador", "Nome Prestador", "IM Prestador", "UF Prestador",
    "CNPJ/CPF Tomador", "Nome Tomador",
    "Cód. Serviço", "Descrição",
    "Valor Serviço", "Base de Cálculo", "Alíquota ISS %", "Valor ISS",
    "Município Incidência ISS", "Cód. IBGE Incidência",
    "Retenção ISS", "Alíq. ISS DPS",
    "BC PIS/COFINS", "Alíq. PIS %", "Alíq. COFINS %", "Valor PIS", "Valor COFINS",
    "Retenção PIS/COF", "CSRF Retido", "CSLL Retida",
    "Valor IRRF", "Valor INSS", "Valor CSLL",
    "Tot. Trib. Fed.", "Tot. Trib. Est.", "Tot. Trib. Mun.",
    "Valor IBS", "Valor CBS", "Alíq. CBS %", "Vlr Líq. + IBS/CBS",
    "Valor Líquido",
    "Tem Evento", "Tipo Evento", "Data Evento", "Motivo Evento",
]
LARG_NFSE = {
    "Arquivo": 26, "Chave NFS-e": 50, "Número NFS-e": 14,
    "Data Emissão": 13, "Competência": 13, "Situação": 22,
    "Unidade Elecnor": 28,
    "Papel Elecnor": 14,
    "CNPJ Prestador": 22, "Nome Prestador": 34, "IM Prestador": 13, "UF Prestador": 7,
    "CNPJ/CPF Tomador": 22, "Nome Tomador": 28,
    "Cód. Serviço": 13, "Descrição": 48,
    "Valor Serviço": 15, "Base de Cálculo": 15,
    "Alíquota ISS %": 13, "Valor ISS": 13,
    "Município Incidência ISS": 28, "Cód. IBGE Incidência": 18,
    "Retenção ISS": 18, "Alíq. ISS DPS": 13,
    "BC PIS/COFINS": 15, "Alíq. PIS %": 11, "Alíq. COFINS %": 13,
    "Valor PIS": 13, "Valor COFINS": 13,
    "Retenção PIS/COF": 28, "CSRF Retido": 12, "CSLL Retida": 12,
    "Valor IRRF": 13, "Valor INSS": 13, "Valor CSLL": 13,
    "Tot. Trib. Fed.": 15, "Tot. Trib. Est.": 15, "Tot. Trib. Mun.": 15,
    "Valor IBS": 13, "Valor CBS": 13, "Alíq. CBS %": 11, "Vlr Líq. + IBS/CBS": 18,
    "Valor Líquido": 13,
    "Tem Evento": 10, "Tipo Evento": 28, "Data Evento": 13, "Motivo Evento": 28,
}
TOTAIS_NFSE = {
    "Valor Serviço", "Base de Cálculo", "Valor ISS",
    "BC PIS/COFINS", "Valor PIS", "Valor COFINS",
    "Valor IRRF", "Valor INSS", "Valor CSLL",
    "Tot. Trib. Fed.", "Tot. Trib. Est.", "Tot. Trib. Mun.",
    "Valor IBS", "Valor CBS", "Vlr Líq. + IBS/CBS",
    "Valor Líquido",
}

COLS_EVT = [
    "Arquivo", "Chave NFS-e", "Tipo de Evento",
    "Data do Evento", "Data Processamento", "Nº Seq. Evento",
    "CNPJ Autor", "Unidade Elecnor", "Cód. Motivo", "Motivo",
]
LARG_EVT = {
    "Arquivo": 28, "Chave NFS-e": 50, "Tipo de Evento": 28,
    "Data do Evento": 14, "Data Processamento": 18, "Nº Seq. Evento": 12,
    "CNPJ Autor": 22, "Unidade Elecnor": 28,
    "Papel Elecnor": 14, "Cód. Motivo": 12, "Motivo": 34,
}

# ── Excel helpers ───────────────────────────────────────────
def cab(cell, bg, fg):
    cell.font      = Font(name="Arial", bold=True, size=10, color=fg)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = Border(
        bottom=Side(style="thin", color="BFD4E5"),
        right =Side(style="thin", color="BFD4E5"),
    )

def dado(cell, bg, chave=""):
    cell.font   = Font(name="Arial", size=10)
    cell.fill   = PatternFill("solid", start_color=bg)
    cell.border = Border(bottom=Side(style="hair", color="D9E3EC"))
    VALS = {
        "Valor Serviço", "Base de Cálculo", "Valor ISS",
        "BC PIS/COFINS", "Valor PIS", "Valor COFINS",
        "Valor IRRF", "Valor INSS", "Valor CSLL",
        "Tot. Trib. Fed.", "Tot. Trib. Est.", "Tot. Trib. Mun.",
        "Valor IBS", "Valor CBS", "Vlr Líq. + IBS/CBS",
        "Valor Líquido",
    }
    if chave in VALS:
        cell.alignment    = Alignment(horizontal="right", vertical="center")
        cell.number_format = "#,##0.00"
    elif chave in ("Alíquota ISS %", "Alíq. ISS DPS", "Alíq. PIS %", "Alíq. COFINS %", "Alíq. CBS %"):
        cell.alignment    = Alignment(horizontal="right", vertical="center")
        cell.number_format = "0.00"
    elif chave in ("Número NFS-e", "Nº Seq. Evento", "Tem Evento"):
        cell.alignment = Alignment(horizontal="center", vertical="center")
    else:
        cell.alignment = Alignment(vertical="center")

def montar_aba(wb, nome_aba, titulo, registros, cabecalhos, larguras,
               cor_tit_bg, cor_tit_fg, cor_cab_bg, cor_cab_fg, cols_total):
    ws = wb.create_sheet(nome_aba[:31])
    ncols = len(cabecalhos)

    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    ws["A1"] = titulo
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color=cor_tit_fg)
    ws["A1"].fill      = PatternFill("solid", start_color=cor_tit_bg)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    for col, c in enumerate(cabecalhos, 1):
        cab(ws.cell(row=2, column=col, value=c), cor_cab_bg, cor_cab_fg)
    ws.row_dimensions[2].height = 34

    for i, reg in enumerate(registros):
        row = i + 3
        bg  = BRANCO if i % 2 == 0 else CINZA
        for col, chave in enumerate(cabecalhos, 1):
            cell = ws.cell(row=row, column=col, value=reg.get(chave))
            dado(cell, bg, chave)
        ws.row_dimensions[row].height = 20

    total_row = len(registros) + 3
    ws.cell(total_row, 1, "TOTAL").font      = Font(name="Arial", bold=True, size=10, color=BRANCO)
    ws.cell(total_row, 1).fill               = PatternFill("solid", start_color=LARANJA)
    ws.cell(total_row, 1).alignment          = Alignment(horizontal="center", vertical="center")
    for col, chave in enumerate(cabecalhos, 1):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = PatternFill("solid", start_color=LARANJA)
        cell.font = Font(name="Arial", bold=True, size=10, color=BRANCO)
        if chave in cols_total:
            letra = get_column_letter(col)
            cell.value         = f"=SUM({letra}3:{letra}{total_row-1})"
            cell.number_format = "#,##0.00"
            cell.alignment     = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[total_row].height = 24

    for col, chave in enumerate(cabecalhos, 1):
        ws.column_dimensions[get_column_letter(col)].width = larguras.get(chave, 14)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}{len(registros)+2}"

# ── main ────────────────────────────────────────────────────
def gerar_excel(pasta_xmls="xmls", pasta_raiz="excel_competencias"):
    os.makedirs(pasta_raiz, exist_ok=True)

    # 1) Parsear todos os XMLs
    todos_nfse    = []
    todos_eventos = []

    for path in listar_xmls(pasta_xmls, "NFSE"):
        try:
            reg = parse_nfse(path)
            if reg: todos_nfse.append(reg)
        except Exception as e:
            print(f"  Erro NFS-e {path}: {e}")

    for path in listar_xmls(pasta_xmls, "EVENTO"):
        try:
            reg = parse_evento(path)
            if reg: todos_eventos.append(reg)
        except Exception as e:
            print(f"  Erro Evento {path}: {e}")

    # 2) Índice de eventos por chave NFS-e
    eventos_por_chave = defaultdict(list)
    for ev in todos_eventos:
        if ev["_chave_nfse"]:
            eventos_por_chave[ev["_chave_nfse"]].append(ev)

    # 3) Enriquecer NFS-e com eventos vinculados
    for reg in todos_nfse:
        evs = eventos_por_chave.get(reg["_chave"], [])
        if evs:
            reg["Tem Evento"]    = "Sim"
            reg["Tipo Evento"]   = " | ".join(e["Tipo de Evento"] for e in evs)
            reg["Data Evento"]   = " | ".join(e["Data do Evento"] for e in evs)
            reg["Motivo Evento"] = " | ".join(e["Motivo"] for e in evs)
            # Situação derivada dos eventos:
            # Se qualquer evento resultar em "Cancelada", a nota é Cancelada.
            # Se apenas eventos neutros (confirmação, bloqueio etc), mantém Gerada.
            situacoes_eventos = [info_evento(e.get("_tag",""))[1] for e in evs]
            if "Cancelada" in situacoes_eventos:
                reg["Situação"] = "Cancelada"
            else:
                # Mantém a situação já definida pelo cStat (Gerada/Substituta)
                pass
        else:
            reg["Tem Evento"] = "Não"
            # Situação pelo cStat quando não há eventos
            cstat = reg.get("_cstat","100")
            if cstat == "101":
                reg["Situação"] = "Substituta"
            elif reg.get("_chsubstda"):
                reg["Situação"] = "Substituta"
            else:
                reg["Situação"] = reg.get("Situação", "Gerada")

    # 4) Agrupar por (unidade, competência)
    grupos_nfse    = defaultdict(lambda: {"tomadas": [], "prestadas": [], "outros": []})
    grupos_eventos = defaultdict(list)

    for reg in todos_nfse:
        chave = (reg["_unidade"], reg["_comp"])
        papel = reg["_papel"]
        if papel == "Tomadora":
            grupos_nfse[chave]["tomadas"].append(reg)
        elif papel == "Prestadora":
            grupos_nfse[chave]["prestadas"].append(reg)
        else:
            grupos_nfse[chave]["outros"].append(reg)

    for ev in todos_eventos:
        chave = (ev["_unidade"], ev["_comp"])
        grupos_eventos[chave].append(ev)

    todas_chaves = sorted(
        set(list(grupos_nfse.keys()) + list(grupos_eventos.keys())),
        key=lambda x: (x[0], x[1])
    )

    print(f"\nNFS-e: {len(todos_nfse)} | Eventos: {len(todos_eventos)}")
    print(f"Combinações filial/competência: {len(todas_chaves)}\n")

    # 5) Gerar estrutura:
    #    excel_competencias/
    #      Elecnor Brasil - Matriz/
    #        NFS-e_2025-09.xlsx
    #        NFS-e_2025-10.xlsx
    #      Elecnor Brasil - Paracatu/
    #        NFS-e_2025-10.xlsx

    for unidade, comp in todas_chaves:
        label_comp = competencia_label(comp)

        pasta_filial = os.path.join(pasta_raiz, slug_seguro(unidade))
        os.makedirs(pasta_filial, exist_ok=True)

        nome_arquivo = f"NFS-e_{comp}.xlsx"
        caminho      = os.path.join(pasta_filial, nome_arquivo)

        tomadas  = grupos_nfse[(unidade, comp)]["tomadas"]
        prestadas = grupos_nfse[(unidade, comp)]["prestadas"]
        evts     = grupos_eventos[(unidade, comp)]

        wb = Workbook()
        wb.remove(wb.active)

        montar_aba(
            wb, "Serviços Tomados",
            f"NFS-e Tomados — {label_comp} — {unidade}",
            tomadas, COLS_NFSE, LARG_NFSE,
            AZUL, BRANCO, AZUL_CLARO, AZUL, TOTAIS_NFSE,
        )
        montar_aba(
            wb, "Serviços Prestados",
            f"NFS-e Prestados — {label_comp} — {unidade}",
            prestadas, COLS_NFSE, LARG_NFSE,
            VERDE, BRANCO, VERDE_CLAR, VERDE, TOTAIS_NFSE,
        )
        montar_aba(
            wb, "Eventos",
            f"Eventos NFS-e — {label_comp} — {unidade}",
            evts, COLS_EVT, LARG_EVT,
            ROXO, BRANCO, ROXO_CLARO, ROXO, set(),
        )

        wb.save(caminho)
        t = len(tomadas); p = len(prestadas); e = len(evts)
        print(f"  {unidade}/")
        print(f"    {nome_arquivo}  →  {t} tomadas | {p} prestadas | {e} eventos")

    print(f"\nEstrutura gerada em: {pasta_raiz}/")

def slug_seguro(texto):
    invalidos = r'\/:*?"<>|'
    for c in invalidos:
        texto = texto.replace(c, "_")
    return texto.strip()




# ── Exportação CSV consolidado ─────────────────────────────
import csv
from eventos_config import info_evento, EVENTOS

def listar_xmls(pasta_raiz: str, tipo: str) -> list:
    """
    Busca XMLs recursivamente em toda a pasta_raiz.
    Retorna lista ordenada de caminhos.
    """
    return sorted(glob.glob(f"{pasta_raiz}/**/{tipo}_*.xml", recursive=True))


def gerar_csv_consolidado(pasta_xmls="xmls", saida_nfse="dados/nfse_consolidado.csv", saida_eventos="dados/eventos_consolidado.csv"):
    os.makedirs("dados", exist_ok=True)

    arquivos_nfse   = listar_xmls(pasta_xmls, "NFSE")
    arquivos_evento = listar_xmls(pasta_xmls, "EVENTO")

    todos_nfse    = []
    todos_eventos = []

    for path in arquivos_nfse:
        try:
            reg = parse_nfse(path)
            if reg:
                todos_nfse.append(reg)
        except Exception as e:
            print(f"  Erro NFS-e {path}: {e}")

    for path in arquivos_evento:
        try:
            reg = parse_evento(path)
            if reg:
                todos_eventos.append(reg)
        except Exception as e:
            print(f"  Erro Evento {path}: {e}")

    # Enriquecer NFS-e com eventos
    eventos_por_chave = defaultdict(list)
    for ev in todos_eventos:
        if ev["_chave_nfse"]:
            eventos_por_chave[ev["_chave_nfse"]].append(ev)

    for reg in todos_nfse:
        evs = eventos_por_chave.get(reg["_chave"], [])
        if evs:
            reg["Tem Evento"]    = "Sim"
            reg["Tipo Evento"]   = " | ".join(e["Tipo de Evento"] for e in evs)
            reg["Data Evento"]   = " | ".join(e["Data do Evento"] for e in evs)
            reg["Motivo Evento"] = " | ".join(e["Motivo"] for e in evs)
            # Situação derivada dos eventos:
            # Se qualquer evento resultar em "Cancelada", a nota é Cancelada.
            # Se apenas eventos neutros (confirmação, bloqueio etc), mantém Gerada.
            situacoes_eventos = [info_evento(e.get("_tag",""))[1] for e in evs]
            if "Cancelada" in situacoes_eventos:
                reg["Situação"] = "Cancelada"
            else:
                # Mantém a situação já definida pelo cStat (Gerada/Substituta)
                pass
        else:
            reg["Tem Evento"] = "Não"
            # Situação pelo cStat quando não há eventos
            cstat = reg.get("_cstat","100")
            if cstat == "101":
                reg["Situação"] = "Substituta"
            elif reg.get("_chsubstda"):
                reg["Situação"] = "Substituta"
            else:
                reg["Situação"] = reg.get("Situação", "Gerada")

    # CSV NFS-e
    colunas_nfse = [c for c in COLS_NFSE if not c.startswith("_")]
    with open(saida_nfse, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=colunas_nfse, extrasaction="ignore")
        writer.writeheader()
        for reg in todos_nfse:
            writer.writerow({k: reg.get(k, "") for k in colunas_nfse})

    # CSV Eventos
    colunas_evt = [c for c in COLS_EVT if not c.startswith("_")]
    with open(saida_eventos, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=colunas_evt, extrasaction="ignore")
        writer.writeheader()
        for reg in todos_eventos:
            writer.writerow({k: reg.get(k, "") for k in colunas_evt})

    print(f"CSV gerado: {saida_nfse} ({len(todos_nfse)} NFS-e)")
    print(f"CSV gerado: {saida_eventos} ({len(todos_eventos)} eventos)")

if __name__ == "__main__":
    gerar_excel()
    gerar_csv_consolidado()
